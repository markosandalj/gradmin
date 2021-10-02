from mature.models import MaturaSubject
from problems.models import AnswerChoice, CorrectAnswer, Problem, Question, Section, Matura, Subject, Term, Video, Year
from openpyxl import load_workbook

# from pathlib import Path 
import pathlib
import json
import requests

from skripte.models import Skripta

headers = {'Content-Type': 'application/json','X-Shopify-Access-Token': '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'}
BASE_PATH = Path("/Volumes/GoogleDrive/My Drive/Projekt instrukcije/Testing ground - Marko/Skripte (ala Trinom)/Fizika/tableFiz.xlsx")

def dbImportSections():
    table = load_workbook(str(BASE_PATH))['Raspored']
    response = requests.get("https://msandalj23.myshopify.com//admin/api/2021-07/pages.json?limit=200", headers=headers)
    response_data = response.json()
    lista_gradiva = []
    lista_gradiva2 = []
    for page in response_data['pages']:
        if(page['template_suffix'] == 'online_skripta_full_view'):
            lista_gradiva.append( (page['title'], page['id']))
    for cell in table['A']:
        page_id = 0
        for page in response_data['pages']:
            if(page['title'] == cell.value):
                page_id = int(page['id'])
        lista_gradiva2.append( (cell.value, page_id) )
        print(page_id, cell.value)
        new_section = Section(
            name=cell.value,
            subject=Subject.objects.get(name='Fizika'),
            shopify_page_id=page_id
        )
        new_section.save()
        new_section.skripta.set( [ Skripta.objects.get(subject__name='Fizika') ] )
    print(list(set(lista_gradiva) - set(lista_gradiva2)))
# dbImportSections()

def saveJson():
    response = requests.get("https://msandalj23.myshopify.com//admin/api/2021-07/pages.json?limit=200", headers=headers)
    with open(str(pathlib.Path().resolve()) + '/problems/data.json', 'w', encoding='utf-8') as write_file:
        json.dump(response.json(), write_file, ensure_ascii=True)

# saveJson()

def getMatura(zad):
    matura_godina = int("20" + zad["matura_godina"][:-2])
    matura_rok = zad["matura_rok"]
    try:
        new_matura=Matura.objects.get(
            year=Year.objects.get(year=matura_godina),
            term=Term.objects.get(term=matura_rok),
            subject=Subject.objects.get(name='Matematika'),
        )
    except:
        new_matura = Matura(
            year=Year.objects.get(year=matura_godina),
            term=Term.objects.get(term=matura_rok),
            subject=Subject.objects.get(name='Matematika'),
        )
        new_matura.save()

def getAnswerChoices(ans, q_text):
    new_ans_choice = AnswerChoice(
        choice_text = ans[3:len(ans)],
        question=Question.objects.get(
                question_text=q_text
            ),
    )
    new_ans_choice.save()

def getCorrectAnswer(zad, q_text):
    if(zad['tocan_odgovor'] != ''):
        new_correct_ans = CorrectAnswer(
            answer_text=zad['tocan_odgovor'],
            question=Question.objects.get(
                question_text=q_text
            ),
        )
        new_correct_ans.save()
    else:
        try:
            new_correct_ans = CorrectAnswer(
                answer_choice = AnswerChoice.objects.filter(
                    question=Question.objects.get(question_text=q_text),
                )[int(zad['tocan_odgovor_zaokruzivanje'])-1],
                question=Question.objects.get(
                    question_text=q_text
                )
            )
            new_correct_ans.save()
        except: 
            print('A: ---> ' + q_text)

def getQuestion(zad):
    q_text = zad['tekst_zadatka']
    new_q = Question(
        question_text=q_text
    )
    new_q.save()


def getVideoVimeo(zad):
    matura_godina = int("20" + zad["matura_godina"][:-2])
    matura_rok = zad["matura_rok"]
    matura_br_zad = zad["tekst_zadatka"].split(' ')[0]
    
    if(zad["vimeo_id"] != '' and zad["vimeo_id"].lower() != 'none' ):
        if('.' in zad["vimeo_id"]):
            video_id = int(zad["vimeo_id"][:-2]) 
        else:
            video_id = int(zad["vimeo_id"]) 
            print(video_id)

        new_video=Video(
            name=f'DM Matematika A - {matura_godina}. {matura_rok}, {matura_br_zad}',
            vimeo_id = video_id
        )
        new_video.save()
        new_video=Video.objects.get(
            name=f'DM Matematika A - {matura_godina}. {matura_rok}, {matura_br_zad}'
        )
        return new_video
    else:
        video_id=None
        return None

def dbImportProblems():
    import_problems_list = []
    with open(str(pathlib.Path().resolve()) + '/problems/gradiva.json', 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        for gradivo in data:
            for zad in data[gradivo]:
                # print(zad['tekst_zadatka'][0:12])
                if(zad['matura_godina'].lower() != 'x' and zad['matura_godina'].lower() != 'medicina'):
                    # matura_godina = int("20" + zad["matura_godina"][:-2])
                    # matura_rok = zad["matura_rok"]
                    # matura_br_zad = zad["tekst_zadatka"][0:3]
                    q_text = zad['tekst_zadatka']
                    # getQuestion(zad)
                    # answers = [
                    #     zad['odgovor_a'],
                    #     zad['odgovor_b'],
                    #     zad['odgovor_c'],
                    #     zad['odgovor_d']
                    # ]
                    # for ans in answers:
                    #     if(ans != ''):
                    #         getAnswerChoices(ans, q_text)
                    getCorrectAnswer(zad, q_text)
                    # new_problem = Problem(
                    #     name=f'DM Matematika {matura_br_zad} {matura_godina}. {matura_rok}',
                    #     matura=Matura.objects.get(
                    #         year=Year.objects.get(year=matura_godina),
                    #         term=Term.objects.get(term=matura_rok),
                    #         subject=Subject.objects.get(name='Matematika'),
                    #     ),
                    #     question=Question.objects.get(
                    #         question_text=q_text
                    #     ),
                    #     subject=Subject.objects.get(
                    #         name="Matematika"
                    #     ),
                    #     section=Section.objects.get(
                    #         name=zad["naziv_gradiva"],
                    #     ),
                    #     video_solution=getVideoVimeo(zad),
                    # )
                    # new_problem.save()
                    import_problems_list.append(new_problem)

# dbImportProblems()


def dbProblemsAddN():
    probs = Problem.objects.all()
    for prob in probs:
        prob.number = str(prob.question).split(' ')[0]
        prob.save()


def removeNoneAnswers():
    noneAnswers = CorrectAnswer.objects.filter( answer_choice=None, answer_text=None)


def getMissingCorrectAnswer(zad, q_text):
    if(zad['tocan_odgovor'] != ''):
        pass
    else:
        try:
            new_correct_ans = CorrectAnswer(
                answer_choice = AnswerChoice.objects.filter(
                    question=Question.objects.get(question_text=q_text),
                )[int(zad['tocan_odgovor_zaokruzivanje'])-1],
                question=Question.objects.get(
                    question_text=q_text
                )
            )
            new_correct_ans.save()
        except: 
            print(q_text)

def importMissingCorrectChoices():
    with open(str(pathlib.Path().resolve()) + '/problems/gradiva.json', 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        for gradivo in data:
            for zad in data[gradivo]:
                if(zad['matura_godina'].lower() != 'x' and zad['matura_godina'].lower() != 'medicina'):
                    q_text = zad['tekst_zadatka']
                    getMissingCorrectAnswer(zad, q_text)

# def importMatProblems():
#     table = load_workbook(str(BASE_PATH))['Zadatci']
#     column = table['A']  #stupac s nazivima gradiva

#     for cell, i in column:
#         row = str(i + 1)
#         value = cell.value
#         if(value != '' and value != 'None'):
#             matura_godina = int('20' + table['B'+row])
#             matura_rok = table['C'+row]
#             matura_br_zad = table['A'+row]
#             q_text = table['E'+row]

def importMatProblems():
    with open(str(pathlib.Path().resolve()) + '/problems/gradiva copy.json', 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        for gradivo in data:
            for zad in data[gradivo]:
                if(zad['matura_godina'].lower() != 'x' and zad['matura_godina'].lower() != 'medicina'):
                    matura_godina = int("20" + zad["matura_godina"])
                    matura_rok = zad["matura_rok"]
                    matura_br_zad = zad["tekst_zadatka"].split(' ')[0]
                    q_text = zad['tekst_zadatka']
                    name=f'DM Matematika A - {matura_godina}. {matura_rok}, {matura_br_zad}'
                    print(name)
                    
                    getQuestion(zad)
                    answers = [
                        zad['odgovor_a'],
                        zad['odgovor_b'],
                        zad['odgovor_c'],
                        zad['odgovor_d']
                    ]
                    for ans in answers:
                        if(ans != ''):
                            getAnswerChoices(ans, q_text)
                    getCorrectAnswer(zad, q_text)
                    try:
                        new_matura=Matura.objects.get(
                            year=Year.objects.get(year=matura_godina),
                            term=Term.objects.get(term=matura_rok),
                            subject=Subject.objects.get(name='Matematika'),
                        )
                    except:
                        new_matura = Matura(
                            year=Year.objects.get(year=matura_godina),
                            term=Term.objects.get(term=matura_rok),
                            subject=Subject.objects.get(name='Matematika'),
                        )
                        new_matura.save()
                    section = Section.objects.get(name=zad["naziv_gradiva"]) if zad["naziv_gradiva"] != 'None' else None
                    
                    new_problem = Problem(
                        name=f'DM Matematika {matura_br_zad} {matura_godina}. {matura_rok}',
                        matura=new_matura,
                        question=Question.objects.get(
                            question_text=q_text
                        ),
                        subject=Subject.objects.get(
                            name="Matematika"
                        ),
                        section=section,
                        video_solution=getVideoVimeo(zad),
                    )
                    new_problem.save()


# -> kreirati sve rokove (ljeto, jesen, zima)
# -> kreirati sve godine za mature (2010 - 2021)
# -> kreirati sve predmete (matematika, fizika, informatika, kemija)
# -> kreirati sve matura predmete - spojiti predmete sa razinama (mat a, mat b)
# -> kreirati sve mature
# kreirati skripte (mata a, mat b, fiz, kem, inf)
# kreirati gradiva (fizika i matematika a)
# import zadataka (fizika, matematika a, matematika b)
    # pitanje 
        # podpitanje (self relacija)
        # ponudeni odgovori
            # text
            # slika
        # tocan odgovor (int 1-4 - choice, else text)
            # choice
            # text
            # slika
        # slika
    # video

# sub = 'Informatika'
# for i in range(2010, 2022):
#     for j in range(0,2):
#         if(j%2==0):
#             term = 'jesen'
#         else:
#             term = 'ljeto'
#         mat = Matura(subject=MaturaSubject.objects.get(subject__name=sub),year=Year.objects.get(year=i),term=Term.objects.get(term=term))
#         mat.save()

def getName(zad, subject, level = 0):
    matura_godina = int("20" + zad["matura_godina"])
    matura_rok = zad["matura_rok"]
    matura_br_zad = zad["tekst_zadatka"].split(' ')[0]
    level_label = ' ' + level if level != 0 else ''
    name=f'{subject}{level_label} - {matura_godina}. {matura_rok}, {matura_br_zad}.'

    return name

def getNumber(zad):
    br_zad = zad["tekst_zadatka"].split(' ')[0]

    return br_zad

def getMatura(zad, subject, level = 0):
    matura_godina = int("20" + zad["matura_godina"])
    matura_rok = zad["matura_rok"]
    try:
        new_matura=Matura.objects.get(
            year=Year.objects.get(year=matura_godina),
            term=Term.objects.get(term=matura_rok),
            subject=MaturaSubject.objects.get(subject__name=subject, level = level),
        )
        return new_matura.id
    except:
        new_matura = Matura(
            year=Year.objects.get(year=matura_godina),
            term=Term.objects.get(term=matura_rok),
            subject=MaturaSubject.objects.get(subject__name=subject, level = level),
        )
        new_matura.save()
        return new_matura.id

def getQuestion(zad):
    q_text = zad['tekst_zadatka']
    try:
        new_q = Question.objects.get(
            question_text=q_text
        )
        return int(new_q.id)
    except:
        new_q = Question(
            question_text=q_text
        )
        new_q.save()
        return int(new_q.id)

def getCorrectAnswer(zad):
    if(zad['tocan_odgovor'] != ''):
        try:
            new_correct_ans = CorrectAnswer(
                answer_text=zad['tocan_odgovor'],
                question=Question.objects.get(
                    id = getQuestion(zad)
                ),
            )
            new_correct_ans.save()
        except:
            print('Somethin went wrong.')
    else:
        try:
            new_correct_ans = CorrectAnswer(
                answer_choice = AnswerChoice.objects.filter(
                    question=Question.objects.get(id = getQuestion(zad)),
                )[int(zad['tocan_odgovor_zaokruzivanje'])-1],
                question=Question.objects.get(
                    id = getQuestion(zad)
                )
            )
            new_correct_ans.save()
        except: 
            print( 'A: ---> ' + Question.objects.get(id = getQuestion(zad)).question_text )

def getAnswerChoices(zad):
    answers = [
        zad['odgovor_a'],
        zad['odgovor_b'],
        zad['odgovor_c'],
        zad['odgovor_d']
    ]
    for ans in answers:
        if(ans != ''):
            try:
                new_ans_choice = AnswerChoice(
                    choice_text = ans[3:len(ans)],
                    question=Question.objects.get(
                             id = getQuestion(zad)
                        ),
                )
                new_ans_choice.save()
            except:
                print('Somethin went wrong')

def getSubject(zad, subject):
    print('Subject: ', subject)
    new_subject = Subject.objects.get(
        name=subject,
    )
    return new_subject.id

def getSection(zad):
    new_section = Section.objects.get(
        name = zad["naziv_gradiva"]
    )

    return new_section.id

def getVideoSolution(zad, subject, level):
    try:
        new_video = Video.objcts.get(
            name = getName(zad, subject, level),
            vimeo_id = int(zad['vimeo_id'])
        )
    except:
        new_video = Video(
            name = getName(zad, subject, level),
            vimeo_id = int(zad['vimeo_id'])
        )
        new_video.save()

    return new_video.id

def getSkripta(zad, subject):
    new_skripta = Skripta.objects.get(
        name = subject + ' - skripta za državnu maturu'
    )
    return new_skripta.id

def getProblem(name, number, matura_id, question_id, subject_id, section_id, video_id, skripta_id):
    try: 
        new_problem = Problem.objects.get(
            name = name,
            number = number,
            matura = Matura.objects.get(id = matura_id),
            question = Question.objects.get(id = question_id),
            subject = Subject.objects.get(id = subject_id),
            section = Section.objects.get(id = section_id),
            video_solution = Video.objects.get(id = video_id),
            skripta = Skripta.objects.get(id = skripta_id),
        )
    except:
        new_problem = Problem(
            name = name,
            number = number,
            matura = Matura.objects.get(id = matura_id),
            question = Question.objects.get(id = question_id),
            subject = Subject.objects.get(id = subject_id),
            section = Section.objects.get(id = section_id),
            video_solution = Video.objects.get(id = video_id),
            skripta = Skripta.objects.get(id = skripta_id),
        )
        new_problem.save()
    
def importProblems(subject, filename, level = 0):
    print('Import starting!')
    filepath = str(pathlib.Path().resolve()) + '/problems/'+ filename + '.json'
    with open(filepath, 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
        for gradivo in data:
            for zad in data[gradivo]:
                if(zad['matura_godina'].lower() != 'x' and zad['matura_godina'].lower() != 'medicina'):
                    name = getName(zad, subject, level)
                    number = getNumber(zad)
                    matura_id = getMatura(zad, subject, level)
                    question_id = getQuestion(zad)
                    getAnswerChoices(zad)
                    getCorrectAnswer(zad)
                    subject_id = getSubject(zad, subject)
                    section_id = getSection(zad)
                    video_id = getVideoSolution(zad, subject, level)
                    skripta_id = getSkripta(zad, subject)
                    getProblem(name, number, matura_id, question_id, subject_id, section_id, video_id, skripta_id)
    print('Import ending!')

importProblems('Fizika', 'gradiva_test')